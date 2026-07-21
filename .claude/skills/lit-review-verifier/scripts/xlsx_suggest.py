#!/usr/bin/env python3
"""
Record a reviewer's suggested changes for ONE paper into a review workbook
(<tracker>.review.xlsx) without ever touching the canonical CSV.

The workbook is a faithful copy of the tracker grid (title row, header row,
then one row per paper). For every field the reviewer wants to change, the
cell is set to the SUGGESTED wording, filled green (Excel "Good" style), and
given a comment holding the original text and the reason. Unchanged cells and
all other rows are left exactly as they were, so the green cells are the diff.

Why JSON-in, not CLI flags: field text contains quotes, commas, and newlines
that are fragile to pass as shell arguments.

Usage:
    python xlsx_suggest.py <csv_path> <suggestion_json_path> [--out <xlsx_path>]

If --out is omitted, the workbook is <csv_path with .csv -> .review.xlsx>.
The workbook is created from the CSV on first run and updated in place on
later runs (so a full pass accumulates every paper's suggestions in one file).

Suggestion JSON shape:
{
  "Paper Title": "Exact title as it appears in the CSV",
  "changes": {
    "Quotable moment": {
      "suggested": "\"...\" (p. 7).",
      "why": "Original cited p. 5; the phrase appears on p. 7 in the extract."
    },
    "Year": { "suggested": "2024", "why": "Title page reads 2024, not 2026." }
  }
}

Only columns present in "changes" are marked. A "changes" of {} is allowed
and means "reviewed, nothing to change" (the workbook is still ensured to
exist so a run over every paper leaves a complete artifact).
"""
import argparse
import csv
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
except ModuleNotFoundError:
    print("ERROR: openpyxl is not installed. Run: python -m pip install openpyxl",
          file=sys.stderr)
    sys.exit(1)

COLUMNS = [
    "Paper Title", "Author(s)", "Year", "Journal or Proceedings", "DOI or URL",
    "Sub-theme", "Central claim", "Key concepts or framework", "Who is centered",
    "Value foreclosed (community value left out)", "Relevance to argument",
    "Gap or limitation", "Quotable moment", "Synthesis paragraph",
]

EM_DASH = "—"

# Excel's built-in "Good" conditional style: light green fill, dark green font.
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100")
WRAP = Alignment(wrap_text=True, vertical="top")


def load_csv(csv_path: Path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = list(csv.reader(f))
    return reader[0], reader[1], reader[2:]  # title_row, header_row, data_rows


def build_workbook(title_row, header_row, data_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    ws.append(title_row)
    ws.append(header_row)
    for row in data_rows:
        # Pad/trim each data row to the header width.
        padded = (row + [""] * len(header_row))[:len(header_row)]
        ws.append(padded)
    for col_idx in range(1, len(header_row) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 32
    return wb


def find_row_index(ws, title: str):
    """1-based worksheet row for the paper, matched on Paper Title (col A)."""
    target = title.strip().lower()
    for r in range(3, ws.max_row + 1):
        cell = ws.cell(row=r, column=1).value
        if cell is not None and str(cell).strip().lower() == target:
            return r
    return None


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("csv_path", type=Path)
    parser.add_argument("suggestion_json_path", type=Path)
    parser.add_argument("--out", type=Path, default=None)
    args = parser.parse_args()

    if not args.csv_path.exists():
        print(f"ERROR: {args.csv_path} does not exist.", file=sys.stderr)
        sys.exit(1)

    out_path = args.out or args.csv_path.with_suffix(".review.xlsx")

    payload = json.loads(args.suggestion_json_path.read_text(encoding="utf-8"))
    title = payload.get("Paper Title", "").strip()
    if not title:
        print("ERROR: suggestion JSON must include a non-empty 'Paper Title'.",
              file=sys.stderr)
        sys.exit(1)
    changes = payload.get("changes", {})

    unknown = [c for c in changes if c not in COLUMNS]
    if unknown:
        print(f"ERROR: unknown column(s): {unknown}\nValid columns: {COLUMNS}",
              file=sys.stderr)
        sys.exit(1)

    # Em-dash guard, same discipline as the extractor's csv_writer.py.
    offenders = [c for c, v in changes.items()
                 if isinstance(v.get("suggested"), str) and EM_DASH in v["suggested"]]
    if offenders:
        print(f"ERROR: em dash (—) found in suggested text for: {offenders}. "
              "Rewrite with a comma, colon, period, or 'and', then retry.",
              file=sys.stderr)
        sys.exit(1)

    title_row, header_row, data_rows = load_csv(args.csv_path)

    # Open the accumulating workbook, or build it fresh from the CSV.
    if out_path.exists():
        wb = load_workbook(out_path)
        ws = wb.active
    else:
        wb = build_workbook(title_row, header_row, data_rows)
        ws = wb.active

    ws_row = find_row_index(ws, title)
    if ws_row is None:
        print(f"ERROR: no row titled '{title}' found in the workbook. "
              "Titles must match the CSV exactly.", file=sys.stderr)
        sys.exit(1)

    applied = []
    for col_name, change in changes.items():
        col_idx = header_row.index(col_name) + 1
        cell = ws.cell(row=ws_row, column=col_idx)
        original = "" if cell.value is None else str(cell.value)
        suggested = str(change.get("suggested", ""))
        why = str(change.get("why", "")).strip()

        cell.value = suggested
        cell.fill = GREEN_FILL
        cell.font = GREEN_FONT
        cell.alignment = WRAP
        comment_text = f"Original: {original or '(empty)'}"
        if why:
            comment_text += f"\n\nWhy: {why}"
        cell.comment = Comment(comment_text, "lit-review-verifier")
        applied.append(col_name)

    wb.save(out_path)

    if applied:
        print(f"Recorded {len(applied)} suggestion(s) for '{title}' "
              f"in {out_path}: {applied}")
    else:
        print(f"Reviewed '{title}': no changes suggested. "
              f"Workbook ensured at {out_path}.")


if __name__ == "__main__":
    main()
